import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active


sheet['A1'] = 'kernel_name'
sheet['B1'] = 'gpu_sim_cycle'
sheet['C1'] = 'L1D_total_cache_accesses'
sheet['D1'] = 'L1D_total_cache_misses'
sheet['E1'] = 'L1D_total_cache_pending_hits'
sheet['F1'] = 'L1D_total_cache_reservation_fails'
sheet['G1'] = 'L1D_cache_data_port_util'
sheet['H1'] = 'L1D_cache_fill_port_util'
sheet['I1'] = 'L2_total_cache_accesses'
sheet['J1'] = 'L2_total_cache_misses'
sheet['K1'] = 'L2_total_cache_pending_hits'
sheet['L1'] = 'L2_total_cache_reservation_fails'
NUM_CHANNEL=64
cnt =1;
while True:
	sheet.cell(row=1, column = cnt+12).value = 'n_act'
	cnt = cnt +1
	if cnt == NUM_CHANNEL+1:
		cnt =1
		break

while True:
	sheet.cell(row=1, column = cnt+12+NUM_CHANNEL+1).value = 'n_rd'
	cnt = cnt +1
	if cnt == NUM_CHANNEL+1:
		cnt =1
		break

while True:
	
	sheet.cell(row=1, column = cnt+12+(NUM_CHANNEL+1)*2).value = 'n_write'
	cnt = cnt +1
	if cnt == NUM_CHANNEL+1:
		cnt = 1
		break

while True:
	sheet.cell(row=1, column = cnt+12+(NUM_CHANNEL+1)*3).value = 'n_wr_bk'
	cnt = cnt +1
	if cnt == NUM_CHANNEL+1:
		cnt =1
		break

while True:
	sheet.cell(row=1, column = cnt+12+65+65+65+65).value = 'bw_util'
	cnt = cnt +1
	if cnt == 65:
		cnt = 1
		break
#title

f = open("kernel_name.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 1).value = line[14:]
	cnt = cnt +1
f.close()


f = open("gpu_sim_cycle.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 2).value = int(line[16:])
	cnt = cnt +1
f.close()

f = open("L1D_total_cache_accesses.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 3).value = int(line[28:])
	cnt = cnt +1
f.close()

f = open("L1D_total_cache_misses.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 4).value = int(line[26:])
	cnt = cnt +1
f.close()

f = open("L1D_total_cache_pending_hits.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 5).value = int(line[32:])
	cnt = cnt +1
f.close()

f = open("L1D_total_cache_reservation_fails.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 6).value = int(line[37:])
	cnt = cnt +1
f.close()

f = open("L1D_cache_data_port_util.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 7).value = float(line[28:])
	cnt = cnt +1
f.close()

f = open("L1D_cache_fill_port_util.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 8).value = float(line[28:])
	cnt = cnt +1
f.close()

f = open("L2_total_cache_accesses.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 9).value = int(line[26:])
	cnt = cnt +1
f.close()

f = open("L2_total_cache_misses.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 10).value = int(line[24:])
	cnt = cnt +1
f.close()

f = open("L2_total_cache_pending_hits.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 11).value = int(line[30:])
	cnt = cnt +1
f.close()

f = open("L2_total_cache_reservation_fails.txt",'r');
cnt =2 
while True:
	line = f.readline()
	if not line : break;
	sheet.cell(row = cnt,column = 12).value = int(line[35:])
	cnt = cnt +1
f.close()

#####################################
col = 13
f = open("n_act.txt",'r');
cnt =2 
while True:
	col = 13

	while True:
		if col == 13+NUM_CHANNEL : break;
		line = f.readline()
		if not line : break;
		sheet.cell(row = cnt,column = col).value = int(line[6:])
		col = col + 1
	if not line : break;
	cnt = cnt +1
f.close()


col = 13+NUM_CHANNEL+1
f = open("n_rd.txt",'r');
cnt =2 
while True:
	col = 13+NUM_CHANNEL+1

	while True:
		if col == 13+NUM_CHANNEL+NUM_CHANNEL+1 : break;
		line = f.readline()
		if not line : break;
		sheet.cell(row = cnt,column = col).value = int(line[5:])
		col = col + 1
	if not line : break;
	cnt = cnt +1
f.close()

col = 13+(NUM_CHANNEL+1)*2
f = open("n_write.txt",'r');
cnt =2 
while True:
	col = 13+(NUM_CHANNEL+1)*2

	while True:
		if col == 13+NUM_CHANNEL+(NUM_CHANNEL+1)*2 : break;
		line = f.readline()
		if not line : break;
		sheet.cell(row = cnt,column = col).value = int(line[8:])
		col = col + 1
	if not line : break;
	cnt = cnt +1
f.close()
col = 13+(NUM_CHANNEL+1)*2


f = open("n_wr_bk.txt",'r');
cnt =2 
while True:
	col = 13+(NUM_CHANNEL+1)*3

	while True:
		if col == 13+NUM_CHANNEL+(NUM_CHANNEL+1)*3 : break;
		line = f.readline()
		if not line : break;
		sheet.cell(row = cnt,column = col).value = int(line[8:])
		col = col + 1
	if not line : break;
	cnt = cnt +1
f.close()


col = 13+(NUM_CHANNEL+1)*4
f = open("bw_util.txt",'r');
cnt =2 
while True:
	col = 13+(NUM_CHANNEL+1)*4

	while True:
		if col == 13+NUM_CHANNEL+(NUM_CHANNEL+1)*4 : break;
		line = f.readline()
		if not line : break;
		sheet.cell(row = cnt,column = col).value = line[8:]
		col = col + 1
	if not line : break;
	cnt = cnt +1
f.close()

wb.save('test.xlsx')

